# Importacion de modulos
import os, sys, time
from openpyxl import load_workbook
from openpyxl import Workbook
from fpdf import FPDF
import openpyxl 

#COLORES                                                                                                                   
GL = "\033[96;1m" # Blue aqua
GL = "\033[96;1m" # Blue aqua
BB = "\033[34;1m" # Blue light
YY = "\033[33;1m" # Yellow light
GG = "\033[32;1m" # Green light
WW = "\033[0;1m"  # White light
RR = "\033[31;1m" # Red light
CC = "\033[36;1m" # Cyan light
B = "\033[34m"    # Blue
Y = "\033[33;1m"  # Yellow
G = "\033[32m"    # Green
W = "\033[0;1m"   # White
R = "\033[31m"    # Red
C = "\033[36;1m"  # Cyan
M = "\033[35;1m"  # Morado

# Funciones
# Animaciones de corridas
def sutil(s):
        for c in s + '\n':
                sys.stdout.write(c)
                sys.stdout.flush()
                time.sleep(12. / 150)

def corrida(s):
        for c in s + '\n':
                sys.stdout.write(c)
                sys.stdout.flush()
                time.sleep(3. / 250)

def valletta(s):
        for c in s + '\n':
                sys.stdout.write(c)
                sys.stdout.flush()
                time.sleep(2. / 120)

def saludo(s):
        for c in s + '\n':
                sys.stdout.write(c)
                sys.stdout.flush()
                time.sleep(3. / 100)

def medio(s):
        for c in s + '\n':
                sys.stdout.write(c)
                sys.stdout.flush()
                time.sleep(8. / 200)

def lento(s):
        for c in s + '\n':
                sys.stdout.write(c)
                sys.stdout.flush()
                time.sleep(12. / 180)

def proceso(s):
        for c in s + '\n':
                sys.stdout.write(c)
                sys.stdout.flush()
                time.sleep(15. / 150)

# Cargar el archivo
arc1 = open("db.txt","r")
arc2 = arc1.read()

# Para los nuevos
if arc2 == "J5YD351Y5QE":
	os.system("clear")
	lento(YY +"OHH, Parece que eres nuevo usando este programa...\nPara que funcione correctamente pon a continuacion el nombre de la base de datos, incluyendo la extension")
	os.system("sleep 1")
	lento(WW +"Asi: archivo.xlsx")
	arc = input(B +"Ingresa el nombre del archivo: "+ WW)
	fill = arc
	arc1 = open("db.txt","w")
	arc3 = arc1.write(arc)
	os.system("rm db.txt;touch db.txt")
	arc4 = open("db.txt","r")
	arc5 = arc4.read()
	arc6 = open("db.txt","w")
	arc7 = arc6.write(arc)

# Para los que ya usaron el archivo

else:
	arc1 = open("db.txt","r")
	arc2 = arc1.read()
	fill = arc2

# Bucle solicitud de celdas - OPCION 2
def menu2():
	ltr1 = input(C +"\nInsertar celda: "+ WW)
	ltr1 = ltr1.upper()
	pgsl = hoja[ltr1].value
	if hoja[ltr1].value is None:
		print (RR +"No existen datos en la celda indicada."+ WW +"\n¿Deseas editarlo?\n"+ GG +"1) Si\n2) No")
		rspp = int(input(C +"\nElige una opcion: "+ WW))
		if rspp == 1:
			rsps = input("Ingresa lo que quieras poner\n>>> ")
			hoja[ltr1] = rsps
			cargar.save(fill)

	else:
		# Busqueda de secciones definidas
		var = ltr1[0]
		fl23 = (var+"1")
		fl24 = hoja[fl23].value
		print (GG +"En la celda solicitada existe el siguiente elemento\n"+ BB +">>>"+ WW, pgsl, "-", R+fl24+WW)
	menu2()

# Programacion general
# Introducción
os.system("clear")
tlt ="""██████╗░██████╗░  ░░░░░░  ██╗░░░██╗████████╗░█████╗░
██╔══██╗██╔══██╗  ░░░░░░  ██║░░░██║╚══██╔══╝██╔══██╗
██████╦╝██║░░██║  █████╗  ╚██╗░██╔╝░░░██║░░░███████║
██╔══██╗██║░░██║  ╚════╝  ░╚████╔╝░░░░██║░░░██╔══██║
██████╦╝██████╔╝  ░░░░░░  ░░╚██╔╝░░░░░██║░░░██║░░██║
╚═════╝░╚═════╝░  ░░░░░░  ░░░╚═╝░░░░░░╚═╝░░░╚═╝░░╚═╝                    v1.02"""
print (YY+tlt)
print (R + "\n-", WW + "Programa creado para almacenar información en bases de datos de VTA.", R + "\n-", WW + "Si no hay datos de algun requerimiento pon el numero 0, busca una manera de como\nsaber la información y completa los espacios en blanco.")

print (BB + "¿Qué es lo que quieres hacer?", M + "\n\n1) Ingresar datos de una V8 nueva\n2) Ver o editar celdas\n3) Ver lista principal de V8\n4) Otros\n5) Salir")

dcs1 = int(input(CC + "\nElige una opción: "+ WW))

print ("")
if dcs1 == 1:

# Eleccion numero 1	
# Procesamiento del archivo
	cargar = load_workbook(fill)
	hoja = cargar.active

# Identificacion de contenido existente
	print (YY +"Las siguientes filas ya estan ocupadas."+ WW)
	fila1 = 1
	while fila1 < 1000:
		fila1 = fila1 +1
		cnto0 = (f"A{fila1}")
		db12 = hoja[cnto0].value
		while db12 != None:
			tom2 = cnto0[1]
			cnt22 = (f"B{tom2}")
			db22 = hoja[cnt22].value
			cnt32 = (f"C{tom2}")
			db32 = hoja[cnt32].value
			print ("Fila:", cnto0, "/\ Código:", db12, "| Nombres:", db22, db32)
			break

		else:
			print (GG +"["+ RR +"!"+ GG +"]"+ RR +" Procura no modificar contenido para evitar perdidas.\n")
			break

# Cuestionario
	fln = int(input(C +"Codigo de V8: "+ WW))
	rspn1 = input(C +"Primer nombre de V8: "+ WW)
	rspn2 = input(C +"Segundo nombre de V8: "+ WW)
	rspa1 = input(C +"Primer apellido de V8: "+ WW)
	rspa2 = input(C +"Segundo apellido de V8: "+ WW)
	rspg = input(C +"Genero de V8: "+ WW)
	rspe = int(input(C +"Edad de V8: "+ WW))
	rspdni = int(input(C +"DNI de V8: "+ WW))
	rspdna = int(input(C +"Digito de DNI de V8: "+ WW))
	rspruc = int(input(C +"RUC de V8: "+ WW))
	rspnac = input(C +"Fecha de Nacimiento de V8: "+ WW)
	rspdis = input(C +"Distrito de V8: "+ WW)
	rspdep = input(C +"Departamento de V8: "+ WW)
	rspal = int(input(C +"Estatura de V8: "+ WW))
	rspub = int(input(C +"Ubigeo de V8: "+ WW))
	rspfde = input(C +"Fecha de emision DNI de V8: "+ WW)
	rspec = input(C +"Estado civil de V8: "+ WW)
	rspnum = input(C +"Número Telefónico de V8: "+ WW)
	rspes = int(input(R +"¿Tiene seguro EsSalud?"+ C +"\n\n1) Si\n2) No"+ R +"\nResponder: "+ WW))
	rspsis = int(input(R +"¿Tiene seguro SIS?"+ C +"\n\n1) Si\n2) No"+ R +"\nResponder: "+ WW))
	rspdom = input(C +"Domicilio de V8: "+ WW)
	rspr = input(C +"Razon de estar en la DB: "+ WW)

# Complemento de calculo
	flt = fln + 1

# Asignacion de filas
	flc = (f"0{fln}")
	fl01 = (f"A{flt}")
	fl1 = (f"B{flt}")
	fl2 = (f"C{flt}")
	fl3 = (f"D{flt}")
	fl4 = (f"E{flt}")
	fl5 = (f"F{flt}")
	fl6 = (f"G{flt}")
	fl7 = (f"H{flt}")
	fl8 = (f"I{flt}")
	fl9 = (f"J{flt}")
	fl10 = (f"K{flt}")
	fl11 = (f"L{flt}")
	fl12 = (f"M{flt}")
	fl13 = (f"N{flt}")
	fl14 = (f"O{flt}")
	fl15 = (f"P{flt}")
	fl16 = (f"Q{flt}")
	fl17 = (f"R{flt}")
	fl18 = (f"S{flt}")
	fl19 = (f"T{flt}")
	fl20 = (f"U{flt}")
	fl21 = (f"V{flt}")
	fl22 = (f"W{flt}")

#--- Identificador de aptitud temporal
	if rspe >= 18:
		edadc = ("Mayor de edad")
	else:
		edadc = ("Menor de edad")

#--- Identificador de seguros
#EsSalud
	if rspes == 1:
		afes = ("Activo")
	else:
		afes = ("Caducado")
#SIS
	if rspsis == 1:
		afsis = ("Activo")
	else:
		afsis = ("Caducado")

# Escritura de Valores
	hoja[fl01] = flc
	hoja[fl1] = rspn1
	hoja[fl2] = rspn2
	hoja[fl3] = rspa1
	hoja[fl4] = rspa2
	hoja[fl5] = rspg
	hoja[fl6] = rspe
	hoja[fl7] = edadc
	hoja[fl8] = rspdni
	hoja[fl9] = rspdna
	hoja[fl10] = rspruc
	hoja[fl11] = rspnac
	hoja[fl12] = rspdis
	hoja[fl13] = rspdep
	hoja[fl14] = rspal
	hoja[fl15] = rspub
	hoja[fl16] = rspfde
	hoja[fl17] = rspec
	hoja[fl18] = rspnum
	hoja[fl19] = afes
	hoja[fl20] = afsis
	hoja[fl21] = rspdom
	hoja[fl22] = rspr

# Guardar
	cargar.save(fill)

# Eleccion numero 2
elif dcs1 == 2:

# Procesamiento del archivo
	cargar = load_workbook(fill)
	hoja = cargar.active

# Selección
	print (GG +"["+ CC +"+"+ GG +"]"+ YY +" Por favor, para insertar una celda primero tendras que poner la\ncolumna (letra) y luego la fila (número).")
	print (GG +"["+ CC +"+"+ GG +"]"+ YY +" Para volver al menú pon CTRL + C")
	try:
		ltr1 = input(C +"\nInsertar celda: "+ WW)
	except AttributeError:
		print ("Vuelve a intentarlo")
		menu2()

# Asignacion de celda solicitada
	ltr1 = ltr1.upper()
	pgsl = hoja[ltr1].value

	if hoja[ltr1].value is None:
		print (RR +"No existen datos en la celda indicada."+ WW +"\n¿Deseas editarlo?\n"+ GG +"1) Si\n2) No")
		rspp = int(input(C +"\nElige una opcion: "+ WW))

		if rspp == 1:
			rsps = input("Ingresa lo que quieras poner\n>>> ")
			hoja[ltr1] = rsps
			cargar.save(fill)

		else:
			pass

	else:

# Busqueda de secciones definidas
		var = ltr1[0]
		fl23 = (var+"1")
		fl24 = hoja[fl23].value

# Imprimir celda solicitada
		print (GG +"En la celda solicitada existe el siguiente elemento\n"+ BB +">>>"+ WW, pgsl, "-", R+fl24+WW)
		menu2()


# Guardar
	cargar.save(fill)

# Eleccion número 3
elif dcs1 == 3:

# Procesamiento de archivo
	cargar = load_workbook(fill)
	hoja = cargar.active

# Preparación para identificación de dato existente
	sutil(YY +"Procesando los datos..."+ WW)
	fila = 1
	while fila < 1000:
		fila = fila +1
		cnto = (f"A{fila}")
		db1 = hoja[cnto].value
		while db1 != None:
			tom = cnto[1]
			cnt2 = (f"B{tom}")
			db2 = hoja[cnt2].value
			cnt3 = (f"C{tom}")
			db3 = hoja[cnt3].value
			print ("Código:", db1, "| Nombres:", db2, db3)
			break

		else:
			print (GG+"Esos son todos los datos existentes\n")
			break

# Opcion 4
elif dcs1 == 4:

# Procesamiento de archivo
	cargar = load_workbook(fill)
	hoja = cargar.active

# OPCIONES ADICIONALES
	print (BB + "Existen las siguientes opciones.", M + "\n1) Recopilación instantanea\n2) Limpiar celdas")
	rspsp = int(input(CC +"Escoge una opcion: "+ WW))

# Subalternativa 1 - Recopilacion instantanea	
	if rspsp == 1:

# Solicitud de datos para recopilacion		
		rspre = input(GG +"\n["+ GL+"+" + GG+"]"+ WW +" Pon el codigo de la V8 que requieras."+ GG +"\n["+ GL +"+"+ GG +"]"+ WW +" Incluye un 0 a la respuesta, ejemplo: 01"+ RR +"\n>>> "+ WW)

# Definicion de Variables sobre cada fila
		rspre0 = int(rspre[1])
		rspre1 = rspre0 + 1
		rspreA = (f"A{rspre1}")
		rspreB = (f"B{rspre1}")
		rspreC = (f"C{rspre1}")
		rspreD = (f"D{rspre1}")
		rspreE = (f"E{rspre1}")
		rspreF = (f"F{rspre1}")
		rspreG = (f"G{rspre1}")
		rspreH = (f"H{rspre1}")
		rspreI = (f"I{rspre1}")
		rspreJ = (f"J{rspre1}")
		rspreK = (f"K{rspre1}")
		rspreL = (f"L{rspre1}")
		rspreM = (f"M{rspre1}")
		rspreN = (f"N{rspre1}")
		rspreO = (f"O{rspre1}")
		rspreP = (f"P{rspre1}")
		rspreQ = (f"Q{rspre1}")
		rspreR = (f"R{rspre1}")
		rspreS = (f"S{rspre1}")
		rspreT = (f"T{rspre1}")
		rspreU = (f"U{rspre1}")
		rspreV = (f"V{rspre1}")
		rspreW = (f"W{rspre1}")

# Generación de Variables definiendo valores de cada celda
		reA = hoja[rspreA].value
		reB = hoja[rspreB].value
		reC = hoja[rspreC].value
		reD = hoja[rspreD].value
		reE = hoja[rspreE].value
		reF = hoja[rspreF].value
		reG = hoja[rspreG].value
		reH = hoja[rspreH].value
		reI = hoja[rspreI].value
		reJ = hoja[rspreJ].value
		reK = hoja[rspreK].value
		reL = hoja[rspreL].value
		reM = hoja[rspreM].value
		reN = hoja[rspreN].value
		reO = hoja[rspreO].value
		reP = hoja[rspreP].value
		reQ = hoja[rspreQ].value
		reR = hoja[rspreR].value
		reS = hoja[rspreS].value
		reT = hoja[rspreT].value
		reU = hoja[rspreU].value
		reV = hoja[rspreV].value
		reW = hoja[rspreW].value

# Impresion de valores con gráfica predeterminada
		print (YY +"=========================================\n")
		print ("Información recopilada de", reB)
		print (B +"\nCódigo:", WW+str(reA), B+"\nNombres:", WW+reB, WW+reC, B+"\nApellidos:", WW+reD, reE, B+"\nGenero:", WW+reF, B+"\nEdad:", WW+str(reG), B+"\nAptitud Temporal:", WW+str(reH), B+"\nDNI:", WW+str(reI), B+"\nDigito de DNI:", WW+str(reJ), B+"\nRUC:", WW+str(reK), B+"\nFecha de Nacimiento:", WW+reL, B+"\nDistrito:", WW+reM, B+"\nDepartamento:", WW+reN, B+"\nEstatura", WW+str(reO), B+"\nUbigeo:", WW+ str(reP), B+"\nFecha de emisión:", WW+reQ, B+"\nEstado Civil:", WW+reR, B+"\nNúmero Telefónico:", WW + reS, B+"\nEstado - Seguro EsSalud:", WW+reT, B+"\nEstado - Seguro SIS:", WW+reU, B+"\nDomicilio:", WW+reV, B+"\nRazón de estar en la DB:", WW+reW)
		print (YY +"\n=========================================\n"+ WW)

# Opciones alternales
		print ("¿Que es lo que deseas hacer?")
		print (M + "\n1) Crear archivo PDF")

		dcss = int(input(CC + "\nElige una opción: "+ WW))

# Seleccion de la primer opcion
		if dcss == 1:
# Creacion de las variables, funciones y hojas
			pdf = FPDF()
			pdf.add_page()
			sutil(YY +"Creando archivo...")
			lento(GG +"Archivo creado con exito\n")
			pdf.set_font("Arial",size=12)
# Insercion de datos en PDF
			pdf.cell(200,10,txt="BD-VTA Archivo recopilado",ln=1,align="C")
			pdf.cell(200,10,txt="Informacion | Fase 1",ln=2,align="r")
			pdf.cell(200,10,txt="Código de la victima: "+ reA,ln=3,align="r")
			pdf.cell(200,10,txt="Nombres: "+ reB+" "+reC,ln=4,align="r")
			pdf.cell(200,10,txt="Apellidos: "+ reD+" "+reE,ln=5,align="r")
			pdf.cell(200,10,txt="Genero: "+ reF,ln=6,align="r")
			pdf.cell(200,10,txt="Edad: "+ str(reG),ln=7,align= "r")
			pdf.cell(200,10,txt="Informacion | Fase 2",ln=8,align="r")
			pdf.cell(200,10,txt="Aptitud Temporal: "+ reH,ln=9,align="r")
			pdf.cell(200,10,txt="DNI: "+ str(reI),ln=10,align="r")
			pdf.cell(200,10,txt="Digito de DNI: "+ str(reJ),ln=11,align="r")
			pdf.cell(200,10,txt="RUC: "+ str(reK),ln=12,align="r")
			pdf.cell(200,10,txt="Fecha de Nacimiento: "+ reL,ln=13,align="r")
			pdf.cell(200,10,txt="Distrito: "+ reM,ln=14,align="r")
			pdf.cell(200,10,txt="Departamento: "+ reN,ln=15,align="r")
			pdf.cell(200,10,txt="Informacion | Fase 3",ln=16,align="r")
			pdf.cell(200,10,txt="Estatura: "+ str(reO),ln=17,align="r")
			pdf.cell(200,10,txt="Ubigeo: "+ str(reP),ln=18,align="r")
			pdf.cell(200,10,txt="Fecha de emisión: "+ reQ,ln=19,align="r")
			pdf.cell(200,10,txt="Estado civil: "+ reR,ln=20,align="r")
			pdf.cell(200,10,txt="Número telefónico: "+ reS,ln=21,align="r")
			pdf.cell(200,10,txt="Estado - Seguro EsSalud: "+ reT,ln=22,align="r")
			pdf.cell(200,10,txt="Estado - Seguro SIS: "+ reU,ln=23,align="r")
			pdf.cell(200,10,txt="Razón de estar en la DB: "+ reW,ln=24,align="r")
# Definicion de titulo			
			arnom = (reB+" "+reC+" "+reD+" "+reE+".pdf")
# Salvacion del archivo final
			pdf.output(arnom)

# Subalternativa 2 - limpiar celdas
	elif rspsp == 2:
		print ("Limpiando celdas")
		fila9 = 1
		while fila9 < 8:
			fila9 = fila9 +1
			cnto9 = (f"A{fila9}")
			db19 = hoja[cnto9].value
			while db19 == None:
				vca = ""
				fila10 = (f"B{fila9}")
				fila11 = (f"C{fila9}")
				fila12 = (f"B{fila9}")
				fila13 = (f"C{fila9}")
				fila14 = (f"D{fila9}")
				fila15 = (f"E{fila9}")
				fila16 = (f"F{fila9}")
				fila17 = (f"G{fila9}")
				fila18 = (f"H{fila9}")
				fila19 = (f"I{fila9}")
				fila20 = (f"J{fila9}")
				fila21 = (f"K{fila9}")
				fila22 = (f"L{fila9}")
				fila23 = (f"M{fila9}")
				fila24 = (f"N{fila9}")
				fila25 = (f"O{fila9}")
				fila26 = (f"P{fila9}")
				fila27 = (f"Q{fila9}")	
				fila28 = (f"R{fila9}")	
				fila29 = (f"S{fila9}")
				fila30 = (f"T{fila9}")
				fila31 = (f"U{fila9}")
				fila32 = (f"V{fila9}")
				fila33 = (f"W{fila9}")
				fila34 = (f"X{fila9}")
				fila35 = (f"Y{fila9}")
				fila36 = (f"Z{fila9}")

				conteo = 10
				while(conteo < 37):
					conteo = conteo + 1
					vlor = (f"fila[{conteo}].value")
					print (vlor)
					print (conteo)
					vlor = None
					break
				else:
					print ("A")
		else:
			print (GG+"Celdas limpias.\n")
			cargar.save(fill)

# Subalternativa incorrecta
	else:
		print ("Escogiste una alternativa incorrecta!")

# Opción 5
elif dcs1 == 5:
	sutil("Saliendo del programa...")
	os.system("clear;figlet VALLETTA")

# Opcion Incorrecta
else:
	print ("¡Escogiste una opcion inexistente!")
	os.system("python3 a1.py")
